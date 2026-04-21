"""
Write extraction results to a formatted Excel file.
"""

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# Column definitions: (header_label, df_key, number_format)
COLUMNS = [
    ('Số PXK',          'so_phieu',    None),
    ('Ngày',            'ngay',        None),
    ('Mã CQT',          'ma_cqt',      None),
    ('D/O No',          'do_no',       None),
    ('Lý do xuất kho',  'ly_do',       None),
    ('Phương tiện',     'phuong_tien', None),
    ('Mã hàng',         'ma_hang',     None),
    ('Tên hàng',        'ten_hang',    None),
    ('ĐVT',             'dvt',         None),
    ('Số lượng',        'so_luong',    '#,##0'),
    ('Đơn giá',         'don_gia',     '#,##0.00'),
    ('Thành tiền',      'thanh_tien',  '#,##0'),
    ('File nguồn',      'file_name',   None),
]

HEADER_BG   = '1F4E79'   # dark blue
HEADER_FG   = 'FFFFFF'
ALT_ROW_BG  = 'D6E4F0'   # light blue for even rows


def results_to_dataframe(results: list) -> pd.DataFrame:
    """Flatten extraction results into a DataFrame."""
    rows = []
    for r in results:
        for item in r['items']:
            rows.append({
                'so_phieu':    r['so_phieu'],
                'ngay':        r['ngay'],
                'ma_cqt':      r.get('ma_cqt', ''),
                'do_no':       r.get('do_no', ''),
                'ly_do':       r.get('ly_do', ''),
                'phuong_tien': r.get('phuong_tien', ''),
                'ma_hang':     item['ma_hang'],
                'ten_hang':    item['ten_hang'],
                'dvt':         item['dvt'],
                'so_luong':    item['so_luong'],
                'don_gia':     item['don_gia'],
                'thanh_tien':  item['thanh_tien'],
                'file_name':   r['file_name'],
            })
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[c[1] for c in COLUMNS])


def results_to_excel(results: list) -> bytes:
    """Convert extraction results to a styled Excel file and return as bytes."""
    df = results_to_dataframe(results)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Dữ liệu PXK', index=False,
                    header=False, startrow=1)

        ws = writer.sheets['Dữ liệu PXK']

        # --- Write headers ---
        header_font = Font(color=HEADER_FG, bold=True, name='Arial', size=10)
        header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG,
                                  fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)

        for col_idx, (label, _, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        ws.row_dimensions[1].height = 30

        # --- Style data rows ---
        alt_fill = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG,
                               fill_type='solid')
        data_font = Font(name='Arial', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align   = Alignment(horizontal='left',   vertical='center')
        right_align  = Alignment(horizontal='right',  vertical='center')

        num_fmt_map = {col_idx: fmt
                       for col_idx, (_, _, fmt) in enumerate(COLUMNS, start=1)
                       if fmt}

        for row_idx, row in enumerate(ws.iter_rows(min_row=2,
                                                    max_row=ws.max_row,
                                                    max_col=len(COLUMNS)),
                                      start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, cell in enumerate(row, start=1):
                cell.font = data_font
                if fill:
                    cell.fill = fill
                if col_idx in num_fmt_map:
                    cell.number_format = num_fmt_map[col_idx]
                    cell.alignment = right_align
                elif col_idx in (1, 2, 5):   # Số PXK, Ngày, ĐVT -> center
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

        # --- Auto-column widths ---
        for col_idx, (label, key, _) in enumerate(COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            if not df.empty and key in df.columns:
                max_val_len = df[key].astype(str).str.len().max()
            else:
                max_val_len = 0
            width = max(len(label), int(max_val_len or 0)) + 4
            ws.column_dimensions[col_letter].width = min(width, 45)

        # --- Freeze header row ---
        ws.freeze_panes = 'A2'

        # --- Auto-filter ---
        ws.auto_filter.ref = ws.dimensions

    output.seek(0)
    return output.getvalue()
