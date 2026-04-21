"""
pxk_core_v5.py - Core module cho việc ghép PXK với học máy từ dữ liệu
"""
import io
import os
import re
import tempfile
from collections import defaultdict
from pathlib import Path

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_DEPS = True
except ImportError as e:
    HAS_DEPS = False
    IMPORT_ERROR = str(e)

from pdf_extractor import extract_pxk


# ============================================================================
# UTILITIES
# ============================================================================

def pxk_sort_key(p):
    """Sort key cho số PXK."""
    try:
        return int(p)
    except:
        return 0


def norm_do_no(s):
    """'0093/0094' → {'93','94'} ; '0096' → {'96'}"""
    if not s:
        return set()
    parts = re.split(r'[/\s]+', str(s).strip())
    res = set()
    for p in parts:
        p = p.strip()
        if p.isdigit():
            res.add(str(int(p)))
    return res


def norm_invoice(s):
    """'00096/1C26TAA001' → '96'"""
    if not s:
        return None
    m = re.match(r'^(\d+)', str(s).strip())
    return str(int(m.group(1))) if m else None


def subset_sum_solutions(values, target, max_sols=10):
    """Tìm các tổ hợp subset sum."""
    t = round(target * 100)
    iv = [round(v * 100) for v in values]
    res = []
    def dfs(start, rem, chosen):
        if len(res) >= max_sols:
            return
        if rem == 0:
            res.append(list(chosen))
            return
        for i in range(start, len(iv)):
            if iv[i] <= rem:
                chosen.append(i)
                dfs(i + 1, rem - iv[i], chosen)
                chosen.pop()
    dfs(0, t, [])
    return res


# ============================================================================
# REFERENCE SCORER - Học từ dữ liệu đã điền
# ============================================================================

class ReferenceScorer:
    """
    Học từ các folder dữ liệu đã điền để chấm điểm các ca mơ hồ.
    """
    
    def __init__(self):
        self.examples = []  # List of (ma_hang, invoice, sl, pxk) tuples
        self.pxk_by_invoice = defaultdict(set)
        self.pxk_by_ma_hang = defaultdict(set)
        self.invoice_ma_to_pxk = {}
        self.folder_count = 0
        self.example_count = 0
    
    def load_from_folder(self, folder_path):
        """Load dữ liệu từ một folder chứa file đã điền."""
        folder = Path(folder_path)
        
        # Tìm file đã điền (thường có tên chứa DA_DIEN, ĐÃ ĐIỀN, etc.)
        excel_files = list(folder.glob('*.xlsx'))
        filled_files = [f for f in excel_files 
                       if any(k in f.name.upper() for k in ['DA_DIEN', 'DIEN', 'ĐÃ ĐIỀN', 'DA DIEN'])]
        
        if not filled_files:
            return 0
        
        filled_file = filled_files[0]
        
        try:
            df = pd.read_excel(filled_file)
            
            # Xác định các cột
            col_ma_hang = self._find_column(df, ['Mã hàng', 'ma_hang', 'MÃ HÀNG'])
            col_invoice = self._find_column(df, ['Số hóa đơn', 'hóa đơn', 'invoice', 'Số hóa đơn trên hệ thống'])
            col_sl = self._find_column(df, ['Số lượng', 'SL', 'Số lượng GR'])
            col_pxk = self._find_column(df, ['Số PXK', 'PXK', 'so_pxk'])
            
            if not all([col_ma_hang, col_pxk]):
                return 0
            
            count = 0
            for _, row in df.iterrows():
                ma_hang = str(row.get(col_ma_hang, '')).strip() if col_ma_hang else ''
                invoice = str(row.get(col_invoice, '')).strip() if col_invoice else ''
                sl = row.get(col_sl, 0) if col_sl else 0
                pxk = str(row.get(col_pxk, '')).strip() if col_pxk else ''
                
                # Normalize
                invoice_norm = norm_invoice(invoice)
                try:
                    sl_float = float(sl) if sl else 0
                except:
                    sl_float = 0
                
                if ma_hang and pxk:
                    self.examples.append({
                        'ma_hang': ma_hang,
                        'invoice': invoice_norm,
                        'sl': sl_float,
                        'pxk': pxk,
                    })
                    
                    if invoice_norm:
                        self.pxk_by_invoice[invoice_norm].add(pxk)
                    self.pxk_by_ma_hang[ma_hang].add(pxk)
                    if invoice_norm:
                        key = (invoice_norm, ma_hang, sl_float)
                        self.invoice_ma_to_pxk[key] = pxk
                    
                    count += 1
            
            self.folder_count += 1
            self.example_count += count
            return count
            
        except Exception as e:
            print(f"Error loading {filled_file}: {e}")
            return 0
    
    def load_from_root(self, root_path='.'):
        """Load dữ liệu từ tất cả các folder con."""
        root = Path(root_path)
        
        # Tìm các folder có tên dạng số (VD: 2033-2096, 2144-2172)
        for folder in root.iterdir():
            if folder.is_dir() and re.match(r'\d{4}-\d{4}', folder.name):
                self.load_from_folder(folder)
        
        return self.example_count
    
    def _find_column(self, df, keywords):
        """Tìm tên cột dựa trên keywords."""
        cols = df.columns.tolist()
        for kw in keywords:
            for col in cols:
                if kw.lower() in str(col).lower():
                    return col
        return None
    
    def score_pxk(self, pxk, ma_hang, invoice=None, sl=None):
        """
        Chấm điểm cho một PXK dựa trên lịch sử.
        
        Returns:
            float: Điểm số (càng cao càng phù hợp)
        """
        score = 0.0
        
        # Nếu có invoice match
        if invoice and invoice in self.pxk_by_invoice:
            if pxk in self.pxk_by_invoice[invoice]:
                score += 10.0
        
        # Nếu có mã hàng match
        if ma_hang in self.pxk_by_ma_hang:
            if pxk in self.pxk_by_ma_hang[ma_hang]:
                score += 5.0
        
        # Nếu có cả invoice + ma_hang + sl match
        if invoice and sl is not None:
            key = (invoice, ma_hang, float(sl))
            if key in self.invoice_ma_to_pxk and self.invoice_ma_to_pxk[key] == pxk:
                score += 20.0
        
        return score
    
    def get_best_candidates(self, candidates, ma_hang, invoice=None, sl=None, top_k=3):
        """Lấy các ứng viên tốt nhất dựa trên điểm số."""
        if not self.examples:
            return candidates[:top_k]
        
        scored = [(pxk, self.score_pxk(pxk, ma_hang, invoice, sl)) for pxk in candidates]
        scored.sort(key=lambda x: (-x[1], pxk_sort_key(x[0])))
        
        return [pxk for pxk, score in scored[:top_k]]


# ============================================================================
# PDF EXTRACTION
# ============================================================================

def extract_pdfs_from_files(file_bytes_list):
    """
    Trích xuất dữ liệu từ nhiều file PDF.
    
    Args:
        file_bytes_list: List of (filename, bytes) tuples
        
    Returns:
        (pxk_totals, pxk_dates, pxk_do_no, errors)
    """
    pxk_totals = defaultdict(lambda: defaultdict(float))
    pxk_dates = {}
    pxk_do_no = {}
    errors = []
    
    with tempfile.TemporaryDirectory() as tmpdir:
        for fname, fbytes in file_bytes_list:
            tmp = os.path.join(tmpdir, fname)
            with open(tmp, 'wb') as f:
                f.write(fbytes)
            
            res = extract_pxk(tmp)
            
            if res.get('error'):
                errors.append({'file': fname, 'error': res['error']})
                continue
            
            pxk = str(res.get('so_phieu', ''))
            if not pxk:
                errors.append({'file': fname, 'error': 'Không tìm được số PXK'})
                continue
            
            ngay = res.get('ngay', '')
            if ngay and pxk not in pxk_dates:
                pxk_dates[pxk] = str(ngay)
            
            do_raw = res.get('do_no', '')
            if do_raw:
                pxk_do_no[pxk] = norm_do_no(do_raw)
            
            for item in res.get('items', []):
                ma_hang = item.get('ma_hang', '')
                so_luong = item.get('so_luong', 0)
                if ma_hang:
                    pxk_totals[pxk][ma_hang] += so_luong
    
    return dict(pxk_totals), pxk_dates, pxk_do_no, errors


# ============================================================================
# FORM PROCESSING
# ============================================================================

def read_form_rows_from_bytes(form_bytes, sheet_name=None):
    """
    Đọc các dòng dữ liệu từ file Excel form.
    
    Returns:
        List of dict: [{'row': int, 'idx': int, 'ma_hang': str, 'sl': float, 'inv': str}, ...]
    """
    wb = openpyxl.load_workbook(io.BytesIO(form_bytes))
    
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif 'XUẤT' in wb.sheetnames:
        ws = wb['XUẤT']
    elif 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
    else:
        ws = wb.active
    
    form_rows = []
    for r in range(2, ws.max_row + 1):
        # Bỏ qua dòng trống
        if ws.cell(r, 2).value is None:
            continue
        
        # Đọc các cột chính (dựa trên cấu trúc từ file mẫu)
        # Cột 4: Mã hàng, Cột 5: Số lượng GR, Cột 3: Số hóa đơn
        ma_hang = str(ws.cell(r, 4).value or '').strip()
        sl_val = ws.cell(r, 5).value or 0
        inv_raw = ws.cell(r, 3).value
        
        try:
            sl = float(sl_val) if sl_val else 0
        except:
            sl = 0
        
        inv = norm_invoice(str(inv_raw).strip()) if inv_raw else None
        
        if ma_hang:
            form_rows.append({
                'row': r,
                'idx': len(form_rows),
                'ma_hang': ma_hang,
                'sl': sl,
                'inv': inv,
            })
    
    return form_rows


# ============================================================================
# MATCHING
# ============================================================================

def match_pxk_v5(form_rows, pxk_totals, pxk_do_no, scorer=None):
    """
    Ghép số PXK với form sử dụng nhiều pha và học máy.
    
    Args:
        form_rows: List các dòng form
        pxk_totals: Dict {pxk: {ma_hang: sl}}
        pxk_do_no: Dict {pxk: set(invoices)}
        scorer: ReferenceScorer object (optional)
        
    Returns:
        (result, status_list, note_pxks)
    """
    all_pxks = sorted(pxk_totals.keys(), key=pxk_sort_key)
    n = len(form_rows)
    assigned = [False] * n
    result = [None] * n
    status = ['no_match'] * n
    note_pxks = [[] for _ in range(n)]
    
    # Build lookup tables
    mh_to_idxs = defaultdict(list)
    inv_mh_to_idxs = defaultdict(list)
    
    for fr in form_rows:
        mh_to_idxs[fr['ma_hang']].append(fr['idx'])
        if fr.get('inv'):
            inv_mh_to_idxs[(fr['inv'], fr['ma_hang'])].append(fr['idx'])
    
    # Build invoice -> eligible PXKs mapping
    inv_to_pxks = defaultdict(list)
    for pxk in all_pxks:
        for d in pxk_do_no.get(pxk, set()):
            inv_to_pxks[d].append(pxk)
    
    def get_free(mh, pxk_dos):
        """Lấy các index chưa gán cho mã hàng."""
        all_free = [i for i in mh_to_idxs[mh] if not assigned[i]]
        if pxk_dos:
            filtered = [i for i in all_free if form_rows[i].get('inv') in pxk_dos]
            return filtered if filtered else all_free
        return all_free
    
    # ── Phase 0: Invoice-consecutive greedy ──────────────────────────────
    resolved_p0 = set()
    for inv in sorted(inv_to_pxks.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        pxks_for_inv = [p for p in sorted(inv_to_pxks[inv], key=pxk_sort_key) 
                       if p not in resolved_p0]
        if not pxks_for_inv:
            continue
        
        mh_set = set()
        for pxk in pxks_for_inv:
            mh_set.update(pxk_totals[pxk].keys())
        
        # Kiểm tra cân bằng tổng
        balanced = True
        for mh in mh_set:
            total_target = sum(pxk_totals[pxk].get(mh, 0) for pxk in pxks_for_inv)
            total_rows = sum(form_rows[i]['sl'] for i in inv_mh_to_idxs.get((inv, mh), [])
                           if not assigned[i])
            if abs(round(total_target * 100) - round(total_rows * 100)) > 1:
                balanced = False
                break
        
        if not balanced:
            continue
        
        # Greedy assignment
        inv_plan = {}
        tentative = set()
        greedy_ok = True
        
        for pxk in pxks_for_inv:
            pxk_plan = {}
            pxk_ok = True
            for mh, target in pxk_totals[pxk].items():
                free = [i for i in inv_mh_to_idxs.get((inv, mh), [])
                       if not assigned[i] and i not in tentative]
                acc = 0.0
                batch = []
                for i in free:
                    sl = form_rows[i]['sl']
                    if round((acc + sl) * 100) <= round(target * 100):
                        acc += sl
                        batch.append(i)
                        if abs(round(acc * 100) - round(target * 100)) < 1:
                            break
                if abs(round(acc * 100) - round(target * 100)) > 1:
                    pxk_ok = False
                    break
                pxk_plan[mh] = batch
            
            if not pxk_ok:
                greedy_ok = False
                break
            inv_plan[pxk] = pxk_plan
            for batch in pxk_plan.values():
                tentative.update(batch)
        
        if greedy_ok:
            for pxk, pxk_plan in inv_plan.items():
                for batch in pxk_plan.values():
                    for i in batch:
                        assigned[i] = True
                        result[i] = pxk
                        status[i] = 'auto'
                resolved_p0.add(pxk)
    
    # ── Phase 1: Subset-sum với scoring ──────────────────────────────────
    unresolved = set(all_pxks) - resolved_p0
    
    for _ in range(50):  # Max iterations
        done = 0
        for pxk in sorted(unresolved, key=pxk_sort_key):
            pxk_dos = pxk_do_no.get(pxk, set())
            plan, ok, unique = {}, True, True
            
            for mh, target in pxk_totals[pxk].items():
                free = get_free(mh, pxk_dos)
                sols = subset_sum_solutions([form_rows[i]['sl'] for i in free], target, 2)
                if not sols:
                    ok = False
                    break
                if len(sols) > 1:
                    unique = False
                    # Dùng scorer để chọn solution tốt nhất
                    if scorer:
                        best_sol = None
                        best_score = -1
                        for sol in sols:
                            score = 0
                            for j in sol:
                                idx = free[j]
                                fr = form_rows[idx]
                                score += scorer.score_pxk(pxk, fr['ma_hang'], fr.get('inv'), fr['sl'])
                            if score > best_score:
                                best_score = score
                                best_sol = sol
                        sols = [best_sol] if best_sol else sols[:1]
                    else:
                        break
                plan[mh] = [free[j] for j in sols[0]]
            
            if ok and unique:
                for idxs in plan.values():
                    for i in idxs:
                        assigned[i] = True
                        result[i] = pxk
                        status[i] = 'auto'
                unresolved.discard(pxk)
                done += 1
        
        if done == 0:
            break
    
    # ── Phase 2: Ambiguous fallback với scoring ──────────────────────────
    pxk_cands = {}
    for pxk in sorted(unresolved, key=pxk_sort_key):
        pxk_dos = pxk_do_no.get(pxk, set())
        plan, ok = {}, True
        
        for mh, target in pxk_totals[pxk].items():
            free = get_free(mh, pxk_dos)
            sols = subset_sum_solutions([form_rows[i]['sl'] for i in free], target, 10)
            if not sols:
                ok = False
                break
            plan[mh] = [free[j] for j in sols[0]]
        
        if ok:
            pxk_cands[pxk] = plan
    
    # Ghi nhận các PXK khả dĩ khác
    for pxk, plan in pxk_cands.items():
        for idxs in plan.values():
            for i in idxs:
                if pxk not in note_pxks[i]:
                    note_pxks[i].append(pxk)
    
    # Sắp xếp theo score nếu có scorer
    if scorer:
        for i in range(n):
            if note_pxks[i]:
                fr = form_rows[i]
                scored = [(pxk, scorer.score_pxk(pxk, fr['ma_hang'], fr.get('inv'), fr['sl']))
                         for pxk in note_pxks[i]]
                scored.sort(key=lambda x: (-x[1], pxk_sort_key(x[0])))
                note_pxks[i] = [pxk for pxk, _ in scored]
    
    # Gán ambiguous
    for pxk in sorted(pxk_cands.keys(), key=pxk_sort_key):
        for idxs in pxk_cands[pxk].values():
            for i in idxs:
                if not assigned[i]:
                    assigned[i] = True
                    result[i] = pxk
                    status[i] = 'ambiguous'
    
    return result, status, note_pxks


# ============================================================================
# EXCEL OUTPUT
# ============================================================================

def build_output_excel(wb_bytes, form_rows, result, status_list, note_pxks, pxk_dates):
    """
    Xây dựng file Excel kết quả.
    
    Args:
        wb_bytes: Bytes của file form gốc
        form_rows: List các dòng form
        result: List số PXK đã ghép
        status_list: List trạng thái
        note_pxks: List các PXK khả dĩ khác
        pxk_dates: Dict ngày PXK
        
    Returns:
        bytes: File Excel kết quả
    """
    FILL_GREEN = PatternFill('solid', fgColor='C6EFCE')
    FILL_YELLOW = PatternFill('solid', fgColor='FFEB9C')
    FILL_RED = PatternFill('solid', fgColor='FFC7CE')
    FILL_HDR = PatternFill('solid', fgColor='1F4E79')
    FILL_BLUE = PatternFill('solid', fgColor='BDD7EE')
    FONT_WHITE = Font(bold=True, color='FFFFFF', name='Arial')
    
    tmp = tempfile.mktemp(suffix='.xlsx')
    with open(tmp, 'wb') as f:
        f.write(wb_bytes)
    
    wb = openpyxl.load_workbook(tmp)
    
    # Xác định sheet
    if 'XUẤT' in wb.sheetnames:
        ws = wb['XUẤT']
    elif 'Sheet1' in wb.sheetnames:
        ws = wb['Sheet1']
    else:
        ws = wb.active
    
    # Thêm header cho các cột mới
    for col, val in [
        (7, 'Số PXK (AUTO)'),
        (17, 'Trạng thái'),
        (18, 'PXK khả dĩ khác'),
        (19, 'Ngày PXK'),
        (20, 'Ghi chú')
    ]:
        c = ws.cell(1, col)
        c.value = val
        c.font = FONT_WHITE
        c.fill = FILL_HDR
    
    # Tìm PXK cuối cùng
    assigned_pxks = [r for r in result if r is not None]
    last_pxk = str(max(assigned_pxks, key=pxk_sort_key)) if assigned_pxks else None
    
    # Điền dữ liệu
    for fr in form_rows:
        r, i = fr['row'], fr['idx']
        pv = result[i]
        sv = status_list[i]
        
        fill = FILL_GREEN if sv == 'auto' else FILL_YELLOW if sv == 'ambiguous' else FILL_RED
        label = 'Tự động' if sv == 'auto' else 'Cần kiểm tra' if sv == 'ambiguous' else 'Không khớp'
        
        # Cột 7: Số PXK
        c7 = ws.cell(r, 7)
        c7.value = f'{int(pv):08d}' if pv and str(pv).isdigit() else (pv or '')
        c7.fill = fill
        
        # Cột 17: Trạng thái
        ws.cell(r, 17).value = label
        ws.cell(r, 17).fill = fill
        
        # Cột 18: PXK khả dĩ khác
        cands = sorted([p for p in note_pxks[i] if p != pv], key=pxk_sort_key)
        if cands:
            ws.cell(r, 18).value = ' | '.join(f'{int(p):08d}' if str(p).isdigit() else p for p in cands[:8])
        
        # Cột 19: Ngày PXK
        if pv and pv in pxk_dates:
            ws.cell(r, 19).value = pxk_dates[pv]
        
        # Cột 20: Ghi chú (đánh dấu PXK cuối)
        if pv and last_pxk and pv == last_pxk:
            ngay = pxk_dates.get(last_pxk, '')
            c20 = ws.cell(r, 20)
            c20.value = f'PXK CUỐI CÙNG ({int(last_pxk):08d}{" - " + ngay if ngay else ""})'
            c20.fill = FILL_BLUE
            c20.font = Font(bold=True, name='Arial')
            c20.alignment = Alignment(wrap_text=True)
    
    wb.save(tmp)
    with open(tmp, 'rb') as f:
        data = f.read()
    os.unlink(tmp)
    
    return data


# ============================================================================
# LOAD REFERENCE SCORER
# ============================================================================

def load_reference_scorer(root_path='.'):
    """Load và trả về ReferenceScorer đã được huấn luyện."""
    scorer = ReferenceScorer()
    scorer.load_from_root(root_path)
    return scorer
