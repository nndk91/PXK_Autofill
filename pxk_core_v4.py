import io
import os
import re
import tempfile
from collections import Counter, defaultdict
from functools import lru_cache

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from pdf_extractor import extract_pxk


def get_data_sheet(wb):
    for name in ("XUẤT ", "Sheet1"):
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def pxk_sort_key(p):
    try:
        return int(p)
    except Exception:
        return 0


def norm_do_no(s):
    parts = re.split(r"[/\s]+", str(s).strip())
    res = set()
    for p in parts:
        p = p.strip()
        if p.isdigit():
            res.add(str(int(p)))
    return res


def norm_invoice(s):
    if not s:
        return None
    m = re.match(r"^(\d+)", str(s).strip())
    return str(int(m.group(1))) if m else None


def subset_sum_solutions(values, target, max_sols=10):
    target_cents = round(target * 100)
    int_values = [round(v * 100) for v in values]
    results = []

    def dfs(start, remaining, chosen):
        if len(results) >= max_sols:
            return
        if remaining == 0:
            results.append(list(chosen))
            return
        for idx in range(start, len(int_values)):
            if int_values[idx] <= remaining:
                chosen.append(idx)
                dfs(idx + 1, remaining - int_values[idx], chosen)
                chosen.pop()

    dfs(0, target_cents, [])
    return results


def read_form_rows_from_bytes(wb_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes))
    ws = get_data_sheet(wb)
    return read_form_rows_from_sheet(ws)


def read_form_rows_from_file(path):
    wb = openpyxl.load_workbook(path)
    ws = get_data_sheet(wb)
    return read_form_rows_from_sheet(ws)


def read_form_rows_from_sheet(ws):
    form_rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is None:
            continue
        mh = str(ws.cell(r, 4).value or "").strip()
        sl = float(ws.cell(r, 5).value or 0)
        inv_raw = ws.cell(r, 3).value
        inv = norm_invoice(str(inv_raw).strip()) if inv_raw else None
        if mh:
            form_rows.append(
                {"row": r, "idx": len(form_rows), "ma_hang": mh, "sl": sl, "inv": inv}
            )
    return form_rows


def read_filled_pxk_values(path):
    wb = openpyxl.load_workbook(path)
    ws = get_data_sheet(wb)
    return {
        r: str(ws.cell(r, 7).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 7).value
    }


def extract_pdfs_from_files(file_bytes_list):
    pxk_totals = defaultdict(lambda: defaultdict(float))
    pxk_dates = {}
    pxk_do_no = {}
    errors = []
    with tempfile.TemporaryDirectory() as tmpdir:
        for fname, fbytes in file_bytes_list:
            tmp = os.path.join(tmpdir, fname)
            with open(tmp, "wb") as f:
                f.write(fbytes)
            res = extract_pxk(tmp)
            if res.get("error"):
                errors.append({"file": fname, "lỗi": res["error"]})
                continue
            pxk = str(res.get("so_phieu", ""))
            if not pxk:
                errors.append({"file": fname, "lỗi": "Không tìm được số PXK"})
                continue
            ngay = res.get("ngay", "")
            if ngay and pxk not in pxk_dates:
                pxk_dates[pxk] = str(ngay)
            do_raw = res.get("do_no", "")
            if do_raw:
                pxk_do_no[pxk] = norm_do_no(do_raw)
            for item in res.get("items", []):
                pxk_totals[pxk][item["ma_hang"]] += item["so_luong"]
    return dict(pxk_totals), pxk_dates, pxk_do_no, errors


def extract_folder_pxk_data(folder_path):
    pxk_folder = os.path.join(folder_path, "PXK")
    pxk_totals = defaultdict(lambda: defaultdict(float))
    pxk_dates = {}
    pxk_do_no = {}
    errors = []

    for fname in sorted(os.listdir(pxk_folder)):
        if not fname.lower().endswith(".pdf"):
            continue
        path = os.path.join(pxk_folder, fname)
        try:
            res = extract_pxk(path)
        except Exception as exc:
            errors.append({"file": fname, "error": str(exc)})
            continue

        if res.get("error"):
            errors.append({"file": fname, "error": res["error"]})
            continue

        pxk = str(res.get("so_phieu", ""))
        if not pxk:
            errors.append({"file": fname, "error": "Không tìm được số PXK"})
            continue

        do_raw = res.get("do_no", "")
        if do_raw:
            pxk_do_no[pxk] = norm_do_no(do_raw)
        ngay = res.get("ngay", "")
        if ngay and pxk not in pxk_dates:
            pxk_dates[pxk] = str(ngay)

        for item in res.get("items", []):
            pxk_totals[pxk][item["ma_hang"]] += item["so_luong"]

    return dict(pxk_totals), pxk_dates, pxk_do_no, errors


def detect_labeled_folders(base_dir):
    folders = []
    for name in sorted(os.listdir(base_dir)):
        folder_path = os.path.join(base_dir, name)
        if not os.path.isdir(folder_path):
            continue
        pxk_dir = os.path.join(folder_path, "PXK")
        if not os.path.isdir(pxk_dir):
            continue
        has_empty = False
        has_filled = False
        for fname in os.listdir(folder_path):
            if not fname.lower().endswith(".xlsx"):
                continue
            low = fname.lower()
            if "không có" in low or "khong co" in low:
                has_empty = True
            if "có dữ liệu" in low or "co du lieu" in low:
                has_filled = True
        if has_empty and has_filled:
            folders.append(folder_path)
    return folders


def find_forms_in_folder(folder_path):
    empty_form = None
    filled_form = None
    for fname in os.listdir(folder_path):
        if not fname.lower().endswith(".xlsx"):
            continue
        low = fname.lower()
        path = os.path.join(folder_path, fname)
        if "không có" in low or "khong co" in low:
            empty_form = path
        elif "có dữ liệu" in low or "co du lieu" in low:
            filled_form = path
    return empty_form, filled_form


def position_bucket(index, size):
    if size <= 1:
        return "0.00"
    ratio = index / (size - 1)
    bucket = round(ratio * 4) / 4
    return f"{bucket:.2f}"


class ReferenceScorer:
    def __init__(self):
        self.rank_counters = defaultdict(Counter)
        self.transition_counters = defaultdict(Counter)
        self.example_count = 0
        self.folder_count = 0

    def add_row_example(
        self,
        inv,
        item_code,
        quantity,
        group_index,
        group_size,
        candidate_count,
        rank,
        prev_rank,
    ):
        qty_key = round(float(quantity), 2)
        pos = position_bucket(group_index, group_size)
        keys = [
            ("inv_item_qty_pos", inv, item_code, qty_key, pos, candidate_count),
            ("item_qty_pos", item_code, qty_key, pos, candidate_count),
            ("item_qty", item_code, qty_key, candidate_count),
            ("item_pos", item_code, pos, candidate_count),
            ("item_only", item_code, candidate_count),
        ]
        for key in keys:
            self.rank_counters[key][rank] += 1
        if prev_rank is not None:
            self.transition_counters[(item_code, candidate_count, prev_rank)][rank] += 1
        self.example_count += 1

    def score_rank(
        self,
        inv,
        item_code,
        quantity,
        group_index,
        group_size,
        candidate_count,
        rank,
        prev_rank=None,
    ):
        qty_key = round(float(quantity), 2)
        pos = position_bucket(group_index, group_size)
        score = 0.0
        weighted_keys = [
            (8.0, ("inv_item_qty_pos", inv, item_code, qty_key, pos, candidate_count)),
            (5.0, ("item_qty_pos", item_code, qty_key, pos, candidate_count)),
            (3.0, ("item_qty", item_code, qty_key, candidate_count)),
            (2.0, ("item_pos", item_code, pos, candidate_count)),
            (1.0, ("item_only", item_code, candidate_count)),
        ]
        for weight, key in weighted_keys:
            counter = self.rank_counters.get(key)
            if not counter:
                continue
            total = sum(counter.values())
            score += weight * (counter.get(rank, 0) / total)

        if prev_rank is not None:
            counter = self.transition_counters.get((item_code, candidate_count, prev_rank))
            if counter:
                total = sum(counter.values())
                score += 4.0 * (counter.get(rank, 0) / total)
        return score


@lru_cache(maxsize=4)
def load_reference_scorer(base_dir):
    scorer = ReferenceScorer()
    for folder_path in detect_labeled_folders(base_dir):
        empty_form, filled_form = find_forms_in_folder(folder_path)
        if not empty_form or not filled_form:
            continue
        pxk_totals, _, pxk_do_no, _ = extract_folder_pxk_data(folder_path)
        form_rows = read_form_rows_from_file(empty_form)
        filled_values = read_filled_pxk_values(filled_form)

        group_rows = defaultdict(list)
        inv_to_pxks = defaultdict(list)
        for pxk in sorted(pxk_totals.keys(), key=pxk_sort_key):
            for inv in pxk_do_no.get(pxk, set()):
                inv_to_pxks[inv].append(pxk)

        for fr in form_rows:
            group_rows[(fr.get("inv"), fr["ma_hang"])].append(fr)

        for (inv, item_code), rows in group_rows.items():
            candidate_pxks = []
            if inv:
                candidate_pxks = [
                    pxk
                    for pxk in sorted(inv_to_pxks.get(inv, []), key=pxk_sort_key)
                    if pxk_totals.get(pxk, {}).get(item_code, 0) > 0
                ]
            if not candidate_pxks:
                continue

            prev_rank = None
            for group_index, fr in enumerate(rows):
                expected = filled_values.get(fr["row"])
                if not expected:
                    continue
                normalized = expected.lstrip("0").replace(".0", "")
                try:
                    rank = next(
                        idx
                        for idx, pxk in enumerate(candidate_pxks)
                        if str(pxk).lstrip("0").replace(".0", "") == normalized
                    )
                except StopIteration:
                    continue
                scorer.add_row_example(
                    inv,
                    item_code,
                    fr["sl"],
                    group_index,
                    len(rows),
                    len(candidate_pxks),
                    rank,
                    prev_rank,
                )
                prev_rank = rank
        scorer.folder_count += 1
    return scorer


def calculate_remaining_quantities(form_rows, result, pxk_totals):
    remaining = {
        pxk: {item_code: qty for item_code, qty in items.items()}
        for pxk, items in pxk_totals.items()
    }
    for fr in form_rows:
        pxk = result[fr["idx"]]
        if not pxk:
            continue
        item_code = fr["ma_hang"]
        if item_code in remaining.get(pxk, {}):
            remaining[pxk][item_code] = max(
                0.0, remaining[pxk][item_code] - float(fr["sl"])
            )
    return remaining


def match_pxk_v4(form_rows, pxk_totals, pxk_do_no, scorer=None):
    all_pxks = sorted(pxk_totals.keys(), key=pxk_sort_key)
    n = len(form_rows)
    assigned = [False] * n
    result = [None] * n
    status = ["no_match"] * n
    note_pxks = [[] for _ in range(n)]

    inv_last_pxk = defaultdict(list)
    mh_to_idxs = defaultdict(list)
    inv_mh_to_idxs = defaultdict(list)

    for fr in form_rows:
        mh_to_idxs[fr["ma_hang"]].append(fr["idx"])
        inv = fr.get("inv")
        if inv:
            inv_mh_to_idxs[(inv, fr["ma_hang"])].append(fr["idx"])

    inv_to_pxks = defaultdict(list)
    for pxk in all_pxks:
        for inv in pxk_do_no.get(pxk, set()):
            inv_to_pxks[inv].append(pxk)

    def score_pxk_proximity(pxk, inv):
        if not inv_last_pxk[inv]:
            return 0.0
        try:
            pxk_num = int(pxk)
            min_dist = min(abs(pxk_num - int(prev)) for prev in inv_last_pxk[inv])
            return max(0.0, 12.0 - min_dist / 10.0)
        except Exception:
            return 0.0

    def get_free(mh, pxk_dos, inv=None):
        all_free = [idx for idx in mh_to_idxs[mh] if not assigned[idx]]
        if pxk_dos:
            filtered = [idx for idx in all_free if form_rows[idx].get("inv") in pxk_dos]
            if filtered:
                return filtered
            if inv:
                inv_only = [idx for idx in all_free if form_rows[idx].get("inv") == inv]
                if inv_only:
                    return inv_only
        return all_free

    resolved_p0 = set()
    for inv in sorted(inv_to_pxks.keys(), key=lambda x: int(x) if str(x).isdigit() else 0):
        pxks_for_inv = [
            pxk
            for pxk in sorted(inv_to_pxks[inv], key=pxk_sort_key)
            if pxk not in resolved_p0
        ]
        if not pxks_for_inv:
            continue

        mh_set = set()
        for pxk in pxks_for_inv:
            mh_set.update(pxk_totals[pxk].keys())

        balanced = True
        for mh in mh_set:
            total_target = sum(pxk_totals[pxk].get(mh, 0) for pxk in pxks_for_inv)
            total_rows = sum(
                form_rows[idx]["sl"]
                for idx in inv_mh_to_idxs.get((inv, mh), [])
                if not assigned[idx]
            )
            if abs(round(total_target * 100) - round(total_rows * 100)) > 1:
                balanced = False
                break
        if not balanced:
            continue

        inv_plan = {}
        tentative = set()
        greedy_ok = True

        for pxk in pxks_for_inv:
            pxk_plan = {}
            pxk_ok = True
            for mh, target in pxk_totals[pxk].items():
                free = [
                    idx
                    for idx in inv_mh_to_idxs.get((inv, mh), [])
                    if not assigned[idx] and idx not in tentative
                ]
                acc = 0.0
                batch = []
                for idx in free:
                    sl = form_rows[idx]["sl"]
                    if round((acc + sl) * 100) <= round(target * 100):
                        acc += sl
                        batch.append(idx)
                        if abs(round(acc * 100) - round(target * 100)) < 1:
                            break
                if abs(round(acc * 100) - round(target * 100)) > 1:
                    pxk_ok = False
                    break
                pxk_plan[mh] = batch
            if not pxk_ok:
                greedy_ok = False
                break
            inv_plan[pxk] = pxk_plan
            for batch in pxk_plan.values():
                tentative.update(batch)

        if greedy_ok:
            for pxk, pxk_plan in inv_plan.items():
                for batch in pxk_plan.values():
                    for idx in batch:
                        assigned[idx] = True
                        result[idx] = pxk
                        status[idx] = "auto"
                        inv_last_pxk[inv].append(pxk)
                        inv_last_pxk[inv] = inv_last_pxk[inv][-5:]
                resolved_p0.add(pxk)

    unresolved = set(all_pxks) - resolved_p0

    for inv in sorted(inv_to_pxks.keys(), key=lambda x: int(x) if str(x).isdigit() else 0):
        pxks_for_inv = [
            pxk for pxk in sorted(inv_to_pxks[inv], key=pxk_sort_key) if pxk in unresolved
        ]
        if not pxks_for_inv:
            continue

        mh_to_pxks = defaultdict(list)
        for pxk in pxks_for_inv:
            for mh in pxk_totals[pxk].keys():
                mh_to_pxks[mh].append(pxk)

        for mh, pxks_with_mh in mh_to_pxks.items():
            if len(pxks_with_mh) <= 1:
                continue
            free_rows = [
                idx for idx in inv_mh_to_idxs.get((inv, mh), []) if not assigned[idx]
            ]
            if not free_rows:
                continue

            total_needed = sum(form_rows[idx]["sl"] for idx in free_rows)
            pxk_quantities = [
                (pxk, pxk_totals[pxk].get(mh, 0))
                for pxk in pxks_with_mh
                if pxk_totals[pxk].get(mh, 0) > 0
            ]
            if not pxk_quantities:
                continue

            target_cents = round(total_needed * 100)
            n_pxks = len(pxk_quantities)
            best_combo = None
            best_gap = float("inf")

            for mask in range(1, 1 << n_pxks):
                combo_sum = sum(
                    round(pxk_quantities[i][1] * 100)
                    for i in range(n_pxks)
                    if mask & (1 << i)
                )
                gap = abs(combo_sum - target_cents)
                if combo_sum == target_cents:
                    best_combo = [
                        pxk_quantities[i][0] for i in range(n_pxks) if mask & (1 << i)
                    ]
                    best_gap = 0
                    break
                if gap < best_gap:
                    best_gap = gap

            if not best_combo:
                continue

            remaining_rows = list(free_rows)
            row_assignments = {}
            for pxk in sorted(best_combo, key=pxk_sort_key):
                target = pxk_totals[pxk].get(mh, 0)
                if not remaining_rows:
                    break
                row_values = [form_rows[idx]["sl"] for idx in remaining_rows]
                sols = subset_sum_solutions(row_values, target, 1)
                if not sols:
                    continue
                picked = [remaining_rows[j] for j in sols[0]]
                row_assignments[pxk] = picked
                for idx in picked:
                    remaining_rows.remove(idx)

            for pxk, rows in row_assignments.items():
                for idx in rows:
                    if assigned[idx]:
                        continue
                    assigned[idx] = True
                    result[idx] = pxk
                    status[idx] = "auto"
                    others = [other for other in best_combo if other != pxk]
                    if others:
                        note_pxks[idx] = others[:]
                unresolved.discard(pxk)

    for _ in range(50):
        progress = 0
        for pxk in sorted(unresolved, key=pxk_sort_key):
            pxk_dos = pxk_do_no.get(pxk, set())
            plan = {}
            ok = True
            unique = True
            for mh, target in pxk_totals[pxk].items():
                free = get_free(mh, pxk_dos)
                sols = subset_sum_solutions([form_rows[idx]["sl"] for idx in free], target, 2)
                if not sols:
                    ok = False
                    break
                if len(sols) > 1:
                    unique = False
                    break
                plan[mh] = [free[j] for j in sols[0]]
            if ok and unique:
                for idxs in plan.values():
                    for idx in idxs:
                        assigned[idx] = True
                        result[idx] = pxk
                        status[idx] = "auto"
                        inv = form_rows[idx].get("inv")
                        if inv:
                            inv_last_pxk[inv].append(pxk)
                            inv_last_pxk[inv] = inv_last_pxk[inv][-5:]
                unresolved.discard(pxk)
                progress += 1
        if progress == 0:
            break

    remaining = calculate_remaining_quantities(form_rows, result, pxk_totals)
    unresolved_groups = defaultdict(list)
    for fr in form_rows:
        if not assigned[fr["idx"]]:
            unresolved_groups[(fr.get("inv"), fr["ma_hang"])].append(fr)

    for (inv, mh), rows in sorted(
        unresolved_groups.items(),
        key=lambda item: (
            int(item[0][0]) if item[0][0] and str(item[0][0]).isdigit() else 0,
            item[0][1],
        ),
    ):
        candidate_pxks = [
            pxk
            for pxk in sorted(inv_to_pxks.get(inv, []), key=pxk_sort_key)
            if remaining.get(pxk, {}).get(mh, 0) > 0.009
        ]
        if not candidate_pxks:
            continue

        beam = [
            {
                "score": 0.0,
                "remaining": tuple(
                    round(remaining[pxk].get(mh, 0.0), 2) for pxk in candidate_pxks
                ),
                "assignments": (),
                "prev_rank": None,
            }
        ]

        for pos, fr in enumerate(rows):
            new_beam = []
            row_candidates = []
            for state in beam:
                for rank, pxk in enumerate(candidate_pxks):
                    rem_qty = state["remaining"][rank]
                    if rem_qty + 0.009 < fr["sl"]:
                        continue
                    next_remaining = list(state["remaining"])
                    next_remaining[rank] = round(rem_qty - fr["sl"], 2)
                    learned = (
                        scorer.score_rank(
                            inv,
                            mh,
                            fr["sl"],
                            pos,
                            len(rows),
                            len(candidate_pxks),
                            rank,
                            state["prev_rank"],
                        )
                        if scorer
                        else 0.0
                    )
                    continuity = 0.0
                    if state["prev_rank"] is not None:
                        if state["prev_rank"] == rank:
                            continuity += 6.0
                        continuity += max(0.0, 2.5 - abs(state["prev_rank"] - rank))
                    fit_bonus = 8.0 if abs(next_remaining[rank]) < 0.01 else max(
                        0.0, 4.0 - next_remaining[rank] / 100.0
                    )
                    proximity = score_pxk_proximity(pxk, inv) if inv else 0.0
                    total_score = state["score"] + learned + continuity + fit_bonus + proximity
                    new_beam.append(
                        {
                            "score": total_score,
                            "remaining": tuple(next_remaining),
                            "assignments": state["assignments"] + (rank,),
                            "prev_rank": rank,
                        }
                    )
                    row_candidates.append(pxk)

            unique_row_candidates = sorted(set(row_candidates), key=pxk_sort_key)
            for fr_row in rows:
                if not note_pxks[fr_row["idx"]]:
                    note_pxks[fr_row["idx"]] = unique_row_candidates[:]

            if not new_beam:
                break

            new_beam.sort(
                key=lambda state: (
                    len(state["assignments"]),
                    round(state["score"], 6),
                    -sum(state["remaining"]),
                ),
                reverse=True,
            )
            dedup = {}
            for state in new_beam:
                key = (state["assignments"], state["remaining"])
                if key not in dedup:
                    dedup[key] = state
                if len(dedup) >= 60:
                    break
            beam = list(dedup.values())
        else:
            best = max(
                beam,
                key=lambda state: (
                    len(state["assignments"]),
                    round(state["score"], 6),
                    -sum(state["remaining"]),
                ),
            )
            for pos, fr in enumerate(rows):
                rank = best["assignments"][pos]
                pxk = candidate_pxks[rank]
                idx = fr["idx"]
                assigned[idx] = True
                result[idx] = pxk
                status[idx] = "ambiguous"
                remaining[pxk][mh] = round(max(0.0, remaining[pxk][mh] - fr["sl"]), 2)
                if inv:
                    inv_last_pxk[inv].append(pxk)
                    inv_last_pxk[inv] = inv_last_pxk[inv][-5:]

    return result, status, note_pxks


def build_output_excel(wb_bytes, form_rows, result, status_list, note_pxks, pxk_dates):
    fill_green = PatternFill("solid", fgColor="C6EFCE")
    fill_yellow = PatternFill("solid", fgColor="FFEB9C")
    fill_red = PatternFill("solid", fgColor="FFC7CE")
    fill_hdr = PatternFill("solid", fgColor="1F4E79")
    fill_blue = PatternFill("solid", fgColor="BDD7EE")
    font_white = Font(bold=True, color="FFFFFF", name="Arial")

    tmp = tempfile.mktemp(suffix=".xlsx")
    with open(tmp, "wb") as f:
        f.write(wb_bytes)
    wb = openpyxl.load_workbook(tmp)
    ws = get_data_sheet(wb)

    for col, val in [
        (7, "Số PXK (AUTO)"),
        (17, "Trạng thái"),
        (18, "PXK khả dĩ khác"),
        (19, "Ngày PXK"),
        (20, "📌 Ghi chú"),
    ]:
        c = ws.cell(1, col)
        c.value = val
        c.font = font_white
        c.fill = fill_hdr

    assigned_pxks = [pxk for pxk in result if pxk is not None]
    last_pxk = str(max(assigned_pxks, key=pxk_sort_key)) if assigned_pxks else None

    for fr in form_rows:
        row = fr["row"]
        idx = fr["idx"]
        pxk = result[idx]
        state = status_list[idx]
        fill = (
            fill_green
            if state == "auto"
            else fill_yellow
            if state == "ambiguous"
            else fill_red
        )
        label = (
            "✅ Tự động"
            if state == "auto"
            else "🔍 Cần kiểm tra"
            if state == "ambiguous"
            else "❌ Không khớp"
        )

        c7 = ws.cell(row, 7)
        c7.value = f"{int(pxk):08d}" if pxk and str(pxk).isdigit() else (pxk or "")
        c7.fill = fill

        ws.cell(row, 17).value = label
        ws.cell(row, 17).fill = fill

        candidates = sorted([p for p in note_pxks[idx] if p != pxk], key=pxk_sort_key)
        if candidates:
            ws.cell(row, 18).value = " | ".join(
                f"{int(p):08d}" if str(p).isdigit() else p for p in candidates[:8]
            )

        if pxk and pxk in pxk_dates:
            ws.cell(row, 19).value = pxk_dates[pxk]

        if pxk and last_pxk and pxk == last_pxk:
            note = ws.cell(row, 20)
            ngay = pxk_dates.get(last_pxk, "")
            note.value = f"⬆ PXK CUỐI CÙNG ({int(last_pxk):08d}{' - ' + ngay if ngay else ''})"
            note.fill = fill_blue
            note.font = Font(bold=True, name="Arial")
            note.alignment = Alignment(wrap_text=True)

    wb.save(tmp)
    with open(tmp, "rb") as f:
        data = f.read()
    os.unlink(tmp)
    return data
