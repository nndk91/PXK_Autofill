"""
app_v2.py — PXK Manager (All-in-One)
=====================================
1. Upload PDF Phiếu Xuất Kho  →  trích xuất dữ liệu
2. Upload FORM CHƯA NHẬP      →  tự động ghép số PXK
3. Tải về file Excel kết quả  (xanh ✅ / vàng 🔍 / đỏ ❌)

Quy tắc ghép (3 pha):
 Phase 0 — Invoice-consecutive greedy (cân bằng tổng):
   Với mỗi hoá đơn, nếu tổng PXK targets == tổng dòng form → gán tuần tự
   theo thứ tự PXK tăng dần. Chỉ commit khi toàn bộ invoice hợp lệ.
 Phase 1 — Iterative subset-sum (unique solution only)
 Phase 2 — Ambiguous fallback (first candidate)
 - Đánh dấu PXK cuối cùng được sử dụng (cột 20)
"""

import io
import os
import re
import tempfile
from collections import defaultdict

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from pdf_extractor import extract_pxk
from excel_writer import results_to_dataframe, results_to_excel

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PXK Manager – SMC",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
.step-header { font-size:1.05rem; font-weight:700; color:#1F4E79; margin-bottom:4px; }
.tag-green  { background:#C6EFCE; color:#276221; padding:2px 8px; border-radius:4px; font-size:.85rem; }
.tag-yellow { background:#FFEB9C; color:#7d6608; padding:2px 8px; border-radius:4px; font-size:.85rem; }
.tag-red    { background:#FFC7CE; color:#9c1f23; padding:2px 8px; border-radius:4px; font-size:.85rem; }
</style>
""", unsafe_allow_html=True)

st.title("📦 PXK Manager – Trích xuất & Ghép số PXK tự động")
st.caption("Upload PDF → Trích xuất → Ghép số PXK vào FORM CHƯA NHẬP → Tải Excel")

# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.header("⚙️ Các bước thực hiện")
    st.markdown('<div class="step-header">Bước 1 — Upload PDF</div>', unsafe_allow_html=True)
    pdf_files = st.file_uploader(
        "Chọn file PDF Phiếu Xuất Kho",
        type=["pdf"], accept_multiple_files=True, key="pdf_upload",
        help="Có thể chọn nhiều file (Ctrl+Click)"
    )
    if pdf_files:
        st.success(f"✅ {len(pdf_files)} file PDF")

    st.divider()
    st.markdown('<div class="step-header">Bước 2 — Upload FORM CHƯA NHẬP</div>', unsafe_allow_html=True)
    form_file = st.file_uploader(
        "FORM DỮ LIỆU CHƯA NHẬP SỐ PXK.xlsx",
        type=["xlsx"], key="form_upload",
    )
    if form_file:
        st.success(f"✅ {form_file.name}")

    st.divider()
    run_btn = st.button(
        "🚀 Bắt đầu xử lý", type="primary",
        disabled=not (pdf_files and form_file), use_container_width=True,
    )
    st.divider()
    st.caption("💡 Màu sắc kết quả:")
    st.markdown('<span class="tag-green">✅ Tự động</span> — 1 PXK duy nhất khớp', unsafe_allow_html=True)
    st.markdown('<span class="tag-yellow">🔍 Cần kiểm tra</span> — nhiều PXK khả dĩ', unsafe_allow_html=True)
    st.markdown('<span class="tag-red">❌ Không khớp</span> — cần điền thủ công', unsafe_allow_html=True)
    st.caption("📌 Cột 20 đánh dấu PXK cuối cùng được dùng.")


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def get_data_sheet(wb):
    for name in ("XUẤT ", "Sheet1"):
        if name in wb.sheetnames:
            return wb[name]
    return wb.active


def pxk_sort_key(p):
    try: return int(p)
    except: return 0


def norm_do_no(s):
    """'0093/0094' → {'93','94'} ;  '0096' → {'96'}"""
    parts = re.split(r'[/\s]+', str(s).strip())
    res = set()
    for p in parts:
        p = p.strip()
        if p.isdigit():
            res.add(str(int(p)))
    return res


def norm_invoice(s):
    """'00096/1C26TAA001' → '96'"""
    if not s:
        return None
    m = re.match(r'^(\d+)', str(s).strip())
    return str(int(m.group(1))) if m else None


def subset_sum_solutions(values, target, max_sols=10):
    t  = round(target * 100)
    iv = [round(v * 100) for v in values]
    res = []
    def dfs(start, rem, chosen):
        if len(res) >= max_sols: return
        if rem == 0: res.append(list(chosen)); return
        for i in range(start, len(iv)):
            if iv[i] <= rem:
                chosen.append(i); dfs(i+1, rem-iv[i], chosen); chosen.pop()
    dfs(0, t, [])
    return res


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — EXTRACT PDFs
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def extract_pdfs(file_bytes_list):
    """Returns (pxk_totals, pxk_dates, pxk_do_no, errors, raw_results)"""
    pxk_totals  = defaultdict(lambda: defaultdict(float))
    pxk_dates   = {}
    pxk_do_no   = {}   # pxk -> set of normalized D/O No strings
    errors      = []
    raw_results = []
    with tempfile.TemporaryDirectory() as tmpdir:
        for fname, fbytes in file_bytes_list:
            tmp = os.path.join(tmpdir, fname)
            with open(tmp, "wb") as f:
                f.write(fbytes)
            res = extract_pxk(tmp)
            if res.get("error"):
                errors.append({"file": fname, "lỗi": res["error"]}); continue
            pxk = str(res.get("so_phieu", ""))
            if not pxk:
                errors.append({"file": fname, "lỗi": "Không tìm được số PXK"}); continue
            raw_results.append(res)
            ngay = res.get("ngay", "")
            if ngay and pxk not in pxk_dates:
                pxk_dates[pxk] = str(ngay)
            do_raw = res.get("do_no", "")
            if do_raw:
                pxk_do_no[pxk] = norm_do_no(do_raw)
            for item in res.get("items", []):
                pxk_totals[pxk][item["ma_hang"]] += item["so_luong"]
    return dict(pxk_totals), pxk_dates, pxk_do_no, errors, raw_results


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — MATCHING (with D/O No constraint)
# ══════════════════════════════════════════════════════════════════════════════

def match_pxk(form_rows, pxk_totals, pxk_do_no):
    """
    Ghép số PXK với 4 pha:
    Phase 0 — Invoice-consecutive greedy (chỉ khi tổng cân bằng):
        Với mỗi hoá đơn, kiểm tra tổng target của các PXK == tổng SL các dòng.
        Nếu cân bằng → gán tuần tự theo thứ tự PXK tăng dần.
    Phase 1 — Multi-PXK subset-sum (NEW): Tìm tổ hợp nhiều PXK cho cùng mã hàng
    Phase 2 — Iterative subset-sum (unique solution only)
    Phase 3 — Ambiguous fallback (first candidate)
    """
    all_pxks = sorted(pxk_totals.keys(), key=pxk_sort_key)
    n = len(form_rows)
    assigned  = [False] * n
    result    = [None]  * n
    status    = ["no_match"] * n
    note_pxks = [[] for _ in range(n)]
    split_info = [[] for _ in range(n)]  # NEW: track multiple PXKs per row

    # Build lookup tables
    mh_to_idxs     = defaultdict(list)
    inv_mh_to_idxs = defaultdict(list)
    inv_to_total_sl = defaultdict(float)  # (inv, mh) -> total quantity in form

    for fr in form_rows:
        mh_to_idxs[fr["ma_hang"]].append(fr["idx"])
        if fr.get("inv"):
            inv_mh_to_idxs[(fr["inv"], fr["ma_hang"])].append(fr["idx"])
            inv_to_total_sl[(fr["inv"], fr["ma_hang"])] += fr["sl"]

    # Build invoice → eligible PXKs mapping
    inv_to_pxks = defaultdict(list)
    for pxk in all_pxks:
        for d in pxk_do_no.get(pxk, set()):
            inv_to_pxks[d].append(pxk)

    def get_free(mh, pxk_dos):
        """Unassigned indices for mh, filtered by D/O No if available."""
        all_free = [i for i in mh_to_idxs[mh] if not assigned[i]]
        if pxk_dos:
            filtered = [i for i in all_free if form_rows[i].get("inv") in pxk_dos]
            return filtered if filtered else all_free
        return all_free

    # ── Phase 0: Invoice-consecutive greedy ────────────────────────────────────
    resolved_p0 = set()
    for inv in sorted(inv_to_pxks.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        pxks_for_inv = [p for p in sorted(inv_to_pxks[inv], key=pxk_sort_key)
                        if p not in resolved_p0]
        if not pxks_for_inv:
            continue

        mh_set = set()
        for pxk in pxks_for_inv:
            mh_set.update(pxk_totals[pxk].keys())

        balanced = True
        for mh in mh_set:
            total_target = sum(pxk_totals[pxk].get(mh, 0) for pxk in pxks_for_inv)
            total_rows = sum(form_rows[i]["sl"] for i in inv_mh_to_idxs.get((inv, mh), [])
                           if not assigned[i])
            if abs(round(total_target * 100) - round(total_rows * 100)) > 1:
                balanced = False
                break
        if not balanced:
            continue

        inv_plan = {}
        tentative = set()
        greedy_ok = True

        for pxk in pxks_for_inv:
            pxk_plan = {}
            pxk_ok = True
            for mh, target in pxk_totals[pxk].items():
                free = [i for i in inv_mh_to_idxs.get((inv, mh), [])
                       if not assigned[i] and i not in tentative]
                acc = 0.0
                batch = []
                for i in free:
                    sl = form_rows[i]["sl"]
                    if round((acc + sl) * 100) <= round(target * 100):
                        acc += sl
                        batch.append(i)
                        if abs(round(acc * 100) - round(target * 100)) < 1:
                            break
                if abs(round(acc * 100) - round(target * 100)) > 1:
                    pxk_ok = False
                    break
                pxk_plan[mh] = batch
            if not pxk_ok:
                greedy_ok = False
                break
            inv_plan[pxk] = pxk_plan
            for batch in pxk_plan.values():
                tentative.update(batch)

        if greedy_ok:
            for pxk, pxk_plan in inv_plan.items():
                for batch in pxk_plan.values():
                    for i in batch:
                        assigned[i] = True
                        result[i] = pxk
                        status[i] = "auto"
                resolved_p0.add(pxk)

    # ── Phase 1: Multi-PXK subset-sum (NEW) ──────────────────────────────────────
    # For each (invoice, item code), find combination of PXKs that sum to form total
    unresolved = set(all_pxks) - resolved_p0

    for inv in sorted(inv_to_pxks.keys(), key=lambda x: int(x) if x.isdigit() else 0):
        pxks_for_inv = [p for p in sorted(inv_to_pxks[inv], key=pxk_sort_key)
                       if p in unresolved]
        if not pxks_for_inv:
            continue

        # Group PXKs by item code available
        mh_to_pxks = defaultdict(list)
        for pxk in pxks_for_inv:
            for mh in pxk_totals[pxk].keys():
                mh_to_pxks[mh].append(pxk)

        for mh, pxks_with_mh in mh_to_pxks.items():
            if len(pxks_with_mh) <= 1:
                continue  # Skip single-PXK items (handled in Phase 2)

            # Get unassigned form rows for this (inv, mh)
            free_rows = [i for i in inv_mh_to_idxs.get((inv, mh), [])
                        if not assigned[i]]
            if not free_rows:
                continue

            # Calculate total needed
            total_needed = sum(form_rows[i]["sl"] for i in free_rows)

            # Get available quantities from each PXK
            pxk_quantities = [(pxk, pxk_totals[pxk].get(mh, 0)) for pxk in pxks_with_mh]
            pxk_quantities = [(p, q) for p, q in pxk_quantities if q > 0]

            if not pxk_quantities:
                continue

            # Try to find combination of PXKs that sum to total_needed
            target_cents = round(total_needed * 100)
            n_pxks = len(pxk_quantities)

            # Simple subset sum across PXKs
            best_combo = None
            best_score = float('inf')

            for mask in range(1, 1 << n_pxks):
                combo_sum = sum(round(pxk_quantities[i][1] * 100) for i in range(n_pxks) if mask & (1 << i))
                if combo_sum == target_cents:
                    combo = [pxk_quantities[i][0] for i in range(n_pxks) if mask & (1 << i)]
                    best_combo = combo
                    break
                elif abs(combo_sum - target_cents) < best_score:
                    best_score = abs(combo_sum - target_cents)

            if best_combo and best_score <= 1:
                # Now distribute rows to PXKs
                remaining_rows = list(free_rows)
                row_assignments = {}  # pxk -> [row_indices]

                for pxk in sorted(best_combo, key=pxk_sort_key):
                    pxk_target = pxk_totals[pxk].get(mh, 0)
                    if not remaining_rows:
                        break

                    row_values = [form_rows[i]["sl"] for i in remaining_rows]
                    sols = subset_sum_solutions(row_values, pxk_target, 1)

                    if sols:
                        assigned_rows = [remaining_rows[j] for j in sols[0]]
                        row_assignments[pxk] = assigned_rows
                        for r in assigned_rows:
                            remaining_rows.remove(r)

                # Mark all assigned rows
                for pxk, rows in row_assignments.items():
                    for i in rows:
                        if not assigned[i]:
                            assigned[i] = True
                            result[i] = pxk
                            status[i] = "auto"
                            # Track other PXKs in this combo
                            other_pxks = [p for p in best_combo if p != pxk]
                            if other_pxks:
                                split_info[i] = other_pxks
                                note_pxks[i] = other_pxks
                for pxk in row_assignments.keys():
                    if pxk in unresolved:
                        unresolved.discard(pxk)

    # ── Phase 2: Iterative subset-sum propagation ──────────────────────────────
    for _ in range(50):
        done = 0
        for pxk in sorted(unresolved, key=pxk_sort_key):
            pxk_dos = pxk_do_no.get(pxk, set())
            plan, ok, unique = {}, True, True
            for mh, target in pxk_totals[pxk].items():
                free = get_free(mh, pxk_dos)
                sols = subset_sum_solutions([form_rows[i]["sl"] for i in free], target, 2)
                if not sols:
                    ok = False
                    break
                if len(sols) > 1:
                    unique = False
                    break
                plan[mh] = [free[j] for j in sols[0]]
            if ok and unique:
                for idxs in plan.values():
                    for i in idxs:
                        assigned[i] = True
                        result[i] = pxk
                        status[i] = "auto"
                unresolved.discard(pxk)
                done += 1
        if done == 0:
            break

    # ── Phase 3: Ambiguous fallback ────────────────────────────────────────────
    pxk_cands = {}
    for pxk in sorted(unresolved, key=pxk_sort_key):
        pxk_dos = pxk_do_no.get(pxk, set())
        plan, ok = {}, True
        for mh, target in pxk_totals[pxk].items():
            free = get_free(mh, pxk_dos)
            sols = subset_sum_solutions([form_rows[i]["sl"] for i in free], target, 10)
            if not sols:
                ok = False
                break
            plan[mh] = [free[j] for j in sols[0]]
        if ok:
            pxk_cands[pxk] = plan

    for pxk, plan in pxk_cands.items():
        for idxs in plan.values():
            for i in idxs:
                if pxk not in note_pxks[i]:
                    note_pxks[i].append(pxk)

    for pxk in sorted(pxk_cands.keys(), key=pxk_sort_key):
        for idxs in pxk_cands[pxk].values():
            for i in idxs:
                if not assigned[i]:
                    assigned[i] = True
                    result[i] = pxk
                    status[i] = "ambiguous"

    return result, status, note_pxks


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — BUILD EXCEL
# ══════════════════════════════════════════════════════════════════════════════

def build_output_excel(wb_bytes, form_rows, result, status_list, note_pxks, pxk_dates):
    FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")
    FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
    FILL_RED    = PatternFill("solid", fgColor="FFC7CE")
    FILL_HDR    = PatternFill("solid", fgColor="1F4E79")
    FILL_BLUE   = PatternFill("solid", fgColor="BDD7EE")
    FONT_WHITE  = Font(bold=True, color="FFFFFF", name="Arial")

    tmp = tempfile.mktemp(suffix=".xlsx")
    with open(tmp, "wb") as f: f.write(wb_bytes)
    wb = openpyxl.load_workbook(tmp)
    ws = get_data_sheet(wb)

    for col, val in [(7,"Số PXK (AUTO)"),(17,"Trạng thái"),(18,"PXK khả dĩ khác"),(19,"Ngày PXK"),(20,"📌 Ghi chú")]:
        c = ws.cell(1, col)
        c.value = val; c.font = FONT_WHITE; c.fill = FILL_HDR

    assigned_pxks = [r for r in result if r is not None]
    last_pxk = str(max(assigned_pxks, key=pxk_sort_key)) if assigned_pxks else None

    for fr in form_rows:
        r, i = fr["row"], fr["idx"]
        pv   = result[i]; sv = status_list[i]
        fill  = FILL_GREEN if sv=="auto" else FILL_YELLOW if sv=="ambiguous" else FILL_RED
        label = "✅ Tự động" if sv=="auto" else "🔍 Cần kiểm tra" if sv=="ambiguous" else "❌ Không khớp"

        c7 = ws.cell(r, 7)
        c7.value = f"{int(pv):08d}" if pv and str(pv).isdigit() else (pv or "")
        c7.fill  = fill
        ws.cell(r,17).value=label; ws.cell(r,17).fill=fill

        cands = sorted([p for p in note_pxks[i] if p!=pv], key=pxk_sort_key)
        if cands:
            ws.cell(r,18).value = " | ".join(f"{int(p):08d}" if str(p).isdigit() else p for p in cands[:8])

        if pv and pv in pxk_dates:
            ws.cell(r,19).value = pxk_dates[pv]

        if pv and last_pxk and pv==last_pxk:
            ngay = pxk_dates.get(last_pxk,"")
            c20 = ws.cell(r,20)
            c20.value = f"⬆ PXK CUỐI CÙNG ({int(last_pxk):08d}{' - '+ngay if ngay else ''})"
            c20.fill  = FILL_BLUE
            c20.font  = Font(bold=True, name="Arial")
            c20.alignment = Alignment(wrap_text=True)

    wb.save(tmp)
    with open(tmp,"rb") as f: data=f.read()
    os.unlink(tmp)
    return data


# ══════════════════════════════════════════════════════════════════════════════
# MAIN UI
# ══════════════════════════════════════════════════════════════════════════════

if not (pdf_files and form_file):
    c1, c2 = st.columns(2)
    c1.info("👈 **Bước 1:** Upload file PDF Phiếu Xuất Kho từ thanh bên trái")
    c2.info("👈 **Bước 2:** Upload FORM DỮ LIỆU CHƯA NHẬP SỐ PXK.xlsx")
    st.stop()

if run_btn or st.session_state.get("processed"):
    if run_btn:
        st.session_state.pop("processed", None)
        st.session_state.pop("results_cache", None)

    if "results_cache" not in st.session_state:
        with st.status("⏳ Đang xử lý...", expanded=True) as sb:
            fbl = [(f.name, f.read()) for f in pdf_files]
            st.write(f"📄 Đang trích xuất {len(fbl)} file PDF...")
            pxk_totals, pxk_dates, pxk_do_no, pdf_errors, raw_results = extract_pdfs(fbl)
            do_covered = sum(1 for p in pxk_totals if p in pxk_do_no)
            st.write(f"✅ {len(pxk_totals)} PXK "
                     f"({do_covered} có D/O No)"
                     + (f" | ⚠️ {len(pdf_errors)} lỗi" if pdf_errors else ""))

            st.write("📊 Đang đọc FORM CHƯA NHẬP...")
            form_bytes = form_file.read()
            wb_in = openpyxl.load_workbook(io.BytesIO(form_bytes))
            ws_in = get_data_sheet(wb_in)
            st.write(f"   Sheet: **{ws_in.title}**")
            form_rows = []
            for r in range(2, ws_in.max_row+1):
                if ws_in.cell(r,2).value is None: continue
                mh  = str(ws_in.cell(r,4).value or "").strip()
                sl  = float(ws_in.cell(r,5).value or 0)
                inv_raw = ws_in.cell(r,3).value
                inv = norm_invoice(str(inv_raw).strip()) if inv_raw else None
                if mh: form_rows.append({"row":r,"idx":len(form_rows),"ma_hang":mh,"sl":sl,"inv":inv})
            st.write(f"✅ {len(form_rows)} dòng dữ liệu")

            st.write("🔄 Đang ghép số PXK (Phase 0: greedy → Phase 1: subset-sum → Phase 2: fallback)...")
            res, stl, np_ = match_pxk(form_rows, pxk_totals, pxk_do_no)
            na = sum(1 for s in stl if s=="auto")
            nb = sum(1 for s in stl if s=="ambiguous")
            nn = sum(1 for s in stl if s=="no_match")
            st.write(f"✅ Auto: **{na}** | Cần KT: **{nb}** | Không khớp: **{nn}**")

            st.write("💾 Đang tạo Excel kết quả...")
            out_bytes = build_output_excel(form_bytes, form_rows, res, stl, np_, pxk_dates)
            extract_bytes = results_to_excel(raw_results)

            st.session_state["results_cache"] = dict(
                pxk_totals=pxk_totals, pxk_dates=pxk_dates, pxk_do_no=pxk_do_no,
                form_rows=form_rows, result=res, status_list=stl,
                note_pxks=np_, output_bytes=out_bytes, pdf_errors=pdf_errors,
                raw_results=raw_results, extract_bytes=extract_bytes,
            )
            st.session_state["processed"] = True
            sb.update(label="✅ Xử lý hoàn tất!", state="complete")

    cache = st.session_state["results_cache"]
    pxk_totals   = cache["pxk_totals"];  pxk_dates   = cache["pxk_dates"]
    pxk_do_no    = cache["pxk_do_no"];   form_rows   = cache["form_rows"]
    result       = cache["result"];       status_list = cache["status_list"]
    note_pxks    = cache["note_pxks"];   output_bytes = cache["output_bytes"]
    pdf_errors   = cache["pdf_errors"];  raw_results  = cache["raw_results"]
    extract_bytes = cache["extract_bytes"]

    n_auto = sum(1 for s in status_list if s=="auto")
    n_amb  = sum(1 for s in status_list if s=="ambiguous")
    n_none = sum(1 for s in status_list if s=="no_match")
    total  = len(form_rows)

    # KPIs
    st.subheader("📊 Kết quả tổng quan")
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("📄 PDF extract",    len(pxk_totals),
              delta=f"-{len(pdf_errors)} lỗi" if pdf_errors else None, delta_color="inverse")
    c2.metric("📋 Dòng form",      total)
    c3.metric("✅ Tự động",         n_auto)
    c4.metric("🔍 Cần kiểm tra",   n_amb)
    c5.metric("❌ Không khớp",      n_none)

    if total > 0:
        pa=n_auto/total*100; pb=n_amb/total*100; pc=n_none/total*100
        st.markdown(f"""
        <div style='background:#FFC7CE;border-radius:6px;height:18px;width:100%;overflow:hidden'>
        <div style='background:#C6EFCE;width:{pa:.1f}%;height:100%;float:left'></div>
        <div style='background:#FFEB9C;width:{pb:.1f}%;height:100%;float:left'></div>
        </div>
        <small>✅ {pa:.1f}% tự động &nbsp;|&nbsp; 🔍 {pb:.1f}% cần KT &nbsp;|&nbsp; ❌ {pc:.1f}% không khớp</small>
        """, unsafe_allow_html=True)

    if pdf_errors:
        with st.expander(f"⚠️ {len(pdf_errors)} file PDF có lỗi"):
            st.dataframe(pd.DataFrame(pdf_errors), hide_index=True, use_container_width=True)

    st.divider()

    # Preview table
    st.subheader("📋 Xem trước kết quả")
    rows_d = []
    for fr in form_rows:
        i=fr["idx"]; pv=result[i]; sv=status_list[i]
        cands=sorted([p for p in note_pxks[i] if p!=pv], key=pxk_sort_key)
        rows_d.append({
            "Dòng":            fr["row"],
            "Mã hàng":         fr["ma_hang"],
            "Số lượng GR":     int(fr["sl"]),
            "Số PXK":          f"{int(pv):08d}" if pv and str(pv).isdigit() else (pv or ""),
            "Ngày PXK":        pxk_dates.get(pv,"") if pv else "",
            "Trạng thái":      "✅ Tự động" if sv=="auto" else "🔍 Cần kiểm tra" if sv=="ambiguous" else "❌ Không khớp",
            "PXK khả dĩ khác": " | ".join(f"{int(p):08d}" if str(p).isdigit() else p for p in cands[:5]),
        })
    df = pd.DataFrame(rows_d)

    cf, cs = st.columns([3,2])
    with cf: search  = st.text_input("🔍 Lọc Mã hàng / Số PXK", placeholder="DC97-...")
    with cs: filt_st = st.selectbox("Lọc trạng thái", ["Tất cả","✅ Tự động","🔍 Cần kiểm tra","❌ Không khớp"])

    df_s = df.copy()
    if search:
        df_s = df_s[df_s["Mã hàng"].str.contains(search,case=False,na=False)|df_s["Số PXK"].str.contains(search,na=False)]
    if filt_st!="Tất cả":
        df_s = df_s[df_s["Trạng thái"]==filt_st]

    st.dataframe(df_s, use_container_width=True, hide_index=True, height=400,
        column_config={
            "Dòng":            st.column_config.NumberColumn(width="small"),
            "Số lượng GR":     st.column_config.NumberColumn(width="small",format="%d"),
            "Số PXK":          st.column_config.TextColumn(width="medium"),
            "Ngày PXK":        st.column_config.TextColumn(width="medium"),
            "Trạng thái":      st.column_config.TextColumn(width="medium"),
            "PXK khả dĩ khác": st.column_config.TextColumn(width="large"),
        })
    st.caption(f"Hiển thị {len(df_s):,} / {len(df):,} dòng")

    st.divider()

    # Download
    st.subheader("⬇️ Tải file kết quả")
    tab_form, tab_extract = st.tabs(["📋 Form đã điền PXK", "📄 Dữ liệu trích xuất PDF"])

    with tab_form:
        c_dl, c_info = st.columns([1, 2])
        with c_dl:
            st.download_button(
                label="📥 Tải FORM ĐÃ ĐIỀN PXK (.xlsx)",
                data=output_bytes,
                file_name="FORM_CHƯA_NHẬP_ĐÃ_ĐIỀN_PXK.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True,
            )
        with c_info:
            st.info(
                "**Cột bổ sung trong Excel:**\n"
                "- Cột 7: Số PXK (AUTO) — 🟢/🟡/🔴\n"
                "- Cột 17: Trạng thái\n"
                "- Cột 18: PXK khả dĩ khác\n"
                "- Cột 19: Ngày PXK\n"
                "- Cột 20: 📌 Ghi chú (đánh dấu PXK cuối)"
            )

    with tab_extract:
        # Preview extracted data (similar to PXK_XNK_SMC app)
        ext_df = results_to_dataframe(raw_results)
        if not ext_df.empty:
            total_pxk  = ext_df["so_phieu"].nunique()
            total_rows = len(ext_df)
            total_sl   = ext_df["so_luong"].sum()
            total_tien = ext_df["thanh_tien"].sum()

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("📄 Số PXK", f"{total_pxk}")
            k2.metric("📋 Dòng hàng", f"{total_rows:,}")
            k3.metric("📦 Tổng số lượng", f"{total_sl:,.0f}")
            k4.metric("💰 Tổng thành tiền", f"{total_tien:,.0f} ₫")

            col_search, col_ma = st.columns([3, 2])
            with col_search:
                search_pxk = st.text_input("🔍 Lọc theo Số PXK", placeholder="VD: 1174", key="ext_pxk")
            with col_ma:
                search_ma = st.text_input("🔍 Lọc theo Mã hàng", placeholder="VD: DC97-22471T", key="ext_ma")

            disp_df = ext_df.copy()
            if search_pxk:
                disp_df = disp_df[disp_df["so_phieu"].str.contains(search_pxk, na=False)]
            if search_ma:
                disp_df = disp_df[disp_df["ma_hang"].str.contains(search_ma, case=False, na=False)]

            st.dataframe(
                disp_df.rename(columns={
                    "so_phieu":   "Số PXK",
                    "ngay":       "Ngày",
                    "do_no":      "D/O No",
                    "ma_hang":    "Mã hàng",
                    "ten_hang":   "Tên hàng",
                    "dvt":        "ĐVT",
                    "so_luong":   "Số lượng",
                    "don_gia":    "Đơn giá",
                    "thanh_tien": "Thành tiền",
                    "file_name":  "File nguồn",
                }),
                use_container_width=True, hide_index=True,
            )

        st.download_button(
            label="📥 Tải dữ liệu trích xuất PDF (.xlsx)",
            data=extract_bytes,
            file_name="du_lieu_pxk_trich_xuat.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=False,
        )
        st.caption(f"File Excel chứa đầy đủ thông tin trích xuất từ {len(raw_results)} phiếu PDF.")
